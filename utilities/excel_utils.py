from openpyxl import load_workbook


class ExcelUtils:

    def __init__(self, file_path):

        self.file_path = file_path

        self.workbook = load_workbook(
            file_path
        )

        self.sheet = self.workbook.active

    def get_row_count(self):

        return self.sheet.max_row

    def get_bike_name(self, row):

        return self.sheet.cell(
            row=row,
            column=2
        ).value

    def write_status(
        self,
        row,
        status
    ):

        self.sheet.cell(
            row=row,
            column=3
        ).value = status

    def write_url(
        self,
        row,
        url
    ):

        self.sheet.cell(
            row=row,
            column=4
        ).value = url

    def save(self):

        self.workbook.save(
            self.file_path
        )